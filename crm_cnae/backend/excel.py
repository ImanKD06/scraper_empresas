"""Exporta a Excel con 3 hojas: Empresas, Leads, Resumen."""
import os
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import config, db


def _ensure_dir():
    p = Path(config.EXCEL_DIR)
    p.mkdir(parents=True, exist_ok=True)
    return p


def _hdr(ws, r, c, v, bg="0F1D35", fg="FFFFFF", sz=10):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Calibri", bold=True, color=fg, size=sz)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return cell


def _cel(ws, r, c, v, bg=None, fg="111827", bold=False, al="left"):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Calibri", size=9, color=fg, bold=bold)
    cell.alignment = Alignment(horizontal=al, vertical="center")
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell


TBGS = {"crecimiento": "D1FAE5", "decrecimiento": "FEE2E2", "estable": "FEF3C7"}
TFGS = {"crecimiento": "059669", "decrecimiento": "DC2626", "estable": "D97706"}
EBGS = {
    "nuevo": "F1F5F9", "contactado": "DBEAFE", "interesado": "FEF3C7",
    "negociacion": "F3E8FF", "cliente": "D1FAE5", "descartado": "FEE2E2",
}
EFGS = {
    "nuevo": "6B7280", "contactado": "1D4ED8", "interesado": "D97706",
    "negociacion": "7C3AED", "cliente": "059669", "descartado": "DC2626",
}


def exportar(asignacion_id: int) -> str:
    out = _ensure_dir()
    asig = db.one("SELECT * FROM asignaciones WHERE id=%s", (asignacion_id,))
    emps = db.query(
        "SELECT * FROM empresas WHERE asignacion_id=%s ORDER BY ranking_posicion, id",
        (asignacion_id,)
    )
    leads = db.query("""
        SELECT l.*, e.nombre AS emp_nom, e.provincia, e.facturacion_actual,
               e.tendencia, e.pct_cambio, e.ranking_posicion, e.cnae, e.licita,
               u.nombre AS com_nom, u.apellidos AS com_ap
        FROM leads l
        JOIN empresas e ON l.empresa_id = e.id
        JOIN usuarios u ON l.comercial_id = u.id
        WHERE e.asignacion_id = %s
        ORDER BY l.actualizado_en DESC
    """, (asignacion_id,))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _hoja_empresas(wb, emps, asig)
    _hoja_leads(wb, leads)
    _hoja_resumen(wb, asig, emps, leads)

    fname = f"CRM_{asig['cnae']}_{asig['provincia']}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    fpath = out / fname
    wb.save(str(fpath))
    return str(fpath)


def _hoja_empresas(wb, emps, asig):
    ws = wb.create_sheet("Empresas")
    ws.merge_cells("A1:N1")
    t = ws["A1"]
    t.value = f"EMPRESAS · CNAE {asig['cnae']} · {asig['provincia']} · {datetime.now().strftime('%d/%m/%Y')}"
    t.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="0F1D35")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    cols = [("Rank.",9),("Nombre Empresa",36),("CIF",13),("Provincia",14),
            ("Municipio",14),("Teléfono",16),("Email",26),("Web",22),
            ("Gerente",20),("Empleados",11),("Facturación",14),
            ("Fac. Anterior",14),("Tendencia",13),("% Cambio",11)]
    for ci,(h,w) in enumerate(cols, 1):
        _hdr(ws, 2, ci, h, bg="2563EB")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    for ri, e in enumerate(emps, 3):
        bg = "F9FAFB" if ri % 2 == 0 else "FFFFFF"
        tend = e.get("tendencia") or ""
        _cel(ws, ri, 1, e["ranking_posicion"], bg=bg, al="center")
        _cel(ws, ri, 2, e["nombre"], bg=bg, bold=True)
        _cel(ws, ri, 3, e["cif"], bg=bg, al="center")
        _cel(ws, ri, 4, e["provincia"], bg=bg)
        _cel(ws, ri, 5, e["municipio"], bg=bg)
        _cel(ws, ri, 6, e["telefono"], bg=bg)
        _cel(ws, ri, 7, e["email"], bg=bg, fg="2563EB")
        _cel(ws, ri, 8, e["web"], bg=bg, fg="2563EB")
        _cel(ws, ri, 9, e["gerente"], bg=bg)
        _cel(ws, ri, 10, e["empleados"], bg=bg, al="center")
        _cel(ws, ri, 11, e["facturacion_actual"], bg=bg, al="right")
        _cel(ws, ri, 12, e["facturacion_anterior"], bg=bg, al="right")
        _cel(ws, ri, 13, tend, bg=TBGS.get(tend, bg), fg=TFGS.get(tend, "6B7280"), bold=True, al="center")
        _cel(ws, ri, 14, e["pct_cambio"], bg=bg, al="center")

    ws.freeze_panes = "A3"
    if emps:
        ws.auto_filter.ref = f"A2:N{2+len(emps)}"


def _hoja_leads(wb, leads):
    ws = wb.create_sheet("Leads CRM")
    ws.merge_cells("A1:L1")
    t = ws["A1"]
    t.value = f"LEADS CRM · {datetime.now().strftime('%d/%m/%Y')}"
    t.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="0F1D35")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    cols = [("Empresa",34),("Comercial",18),("Estado",13),("Interés",9),
            ("Contacto",20),("Cargo",16),("Teléfono",16),("Email",24),
            ("Facturación",14),("Tendencia",12),("Notas",36),("Próx. Acción",26)]
    for ci,(h,w) in enumerate(cols, 1):
        _hdr(ws, 2, ci, h, bg="2563EB")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    for ri, l in enumerate(leads, 3):
        bg = "F9FAFB" if ri % 2 == 0 else "FFFFFF"
        est = l["estado"] or "nuevo"
        _cel(ws, ri, 1, l["emp_nom"], bg=bg, bold=True)
        _cel(ws, ri, 2, f"{l['com_nom']} {l['com_ap']}".strip(), bg=bg)
        _cel(ws, ri, 3, est, bg=EBGS.get(est, bg), fg=EFGS.get(est, "6B7280"), bold=True, al="center")
        _cel(ws, ri, 4, "★" * (l["interes"] or 0), bg=bg, al="center")
        _cel(ws, ri, 5, l["nombre_contacto"], bg=bg)
        _cel(ws, ri, 6, l["cargo_contacto"], bg=bg)
        _cel(ws, ri, 7, l["telefono_contacto"], bg=bg)
        _cel(ws, ri, 8, l["email_contacto"], bg=bg, fg="2563EB")
        _cel(ws, ri, 9, l["facturacion_actual"], bg=bg, al="right")
        tend = l["tendencia"] or ""
        _cel(ws, ri, 10, tend, bg=TBGS.get(tend, bg), fg=TFGS.get(tend, "6B7280"), bold=True, al="center")
        _cel(ws, ri, 11, l["notas"], bg=bg)
        _cel(ws, ri, 12, l["proxima_accion"], bg=bg)

    ws.freeze_panes = "A3"


def _hoja_resumen(wb, asig, emps, leads):
    ws = wb.create_sheet("Resumen")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.merge_cells("A1:B1")
    t = ws["A1"]
    t.value = "RESUMEN EJECUTIVO"
    t.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="0F1D35")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    crec  = sum(1 for e in emps if e["tendencia"] == "crecimiento")
    decr  = sum(1 for e in emps if e["tendencia"] == "decrecimiento")
    estab = sum(1 for e in emps if e["tendencia"] == "estable")

    filas = [
        ("BÚSQUEDA", None),
        ("CNAE", asig["cnae"]),
        ("Provincia", asig["provincia"]),
        ("Páginas scrapeadas", asig["paginas"]),
        ("Total empresas", len(emps)),
        ("", ""),
        ("TENDENCIAS", None),
        ("▲ En crecimiento", crec),
        ("▼ En decrecimiento", decr),
        ("— Estables", estab),
        ("", ""),
        ("LEADS CRM", None),
        ("Total leads", len(leads)),
        ("Clientes", sum(1 for l in leads if l["estado"] == "cliente")),
        ("En negociación", sum(1 for l in leads if l["estado"] == "negociacion")),
        ("Interesados", sum(1 for l in leads if l["estado"] == "interesado")),
        ("Contactados", sum(1 for l in leads if l["estado"] == "contactado")),
        ("Descartados", sum(1 for l in leads if l["estado"] == "descartado")),
    ]
    for ri, (k, v) in enumerate(filas, 2):
        if v is None:
            _hdr(ws, ri, 1, k, bg="2563EB")
            ws.merge_cells(f"A{ri}:B{ri}")
        elif k == "":
            ws.row_dimensions[ri].height = 5
        else:
            _cel(ws, ri, 1, k, bg="F9FAFB", bold=True)
            _cel(ws, ri, 2, v)